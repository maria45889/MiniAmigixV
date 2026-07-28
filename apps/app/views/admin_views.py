"""
Admin views.
"""

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from django.utils import timezone
from datetime import timedelta
from django.conf import settings
import logging

logger = logging.getLogger(__name__)


def admin_only(view_func):
    """
    Decorador personalizado que solo permite acceso al email del admin principal.
    """
    def wrapped_view(request, *args, **kwargs):
        allowed_admins = getattr(settings, 'ADMIN_EMAILS', ['miniamigixv@gmail.com'])
        if isinstance(allowed_admins, str):
            allowed_admins = [allowed_admins]
        allowed_admins = [email.strip().lower() for email in allowed_admins if email]
        
        user_email = (getattr(request.user, 'email', '') or '').strip().lower()
        
        if not (request.user and request.user.is_authenticated and user_email in allowed_admins):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("No tienes permiso para acceder a esta página.")
        
        return view_func(request, *args, **kwargs)
    return wrapped_view


@admin_only
def panel_admin(request):
    """Render admin panel dashboard with statistics."""
    from django.db.models import Count, Q, Avg
    from apps.app.models import ConversacionChat, Cancion, Playlist, Favorite
    from apps.eventos.models import Evento
    from apps.notificaciones.models import Notificacion
    from apps.soporte.models import TicketSoporte
    from apps.blog.models import Post, Category, Comment
    from apps.sugerencias.models import Sugerencia, Visitante
    from apps.estudio.models import StudyResource, StudySession, PomodoroSession
    from apps.clima.models import WeatherCache
    from apps.traductor.models import TranslationCache
    
    # Calculate dates for filtering
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # User statistics
    total_usuarios = User.objects.count()
    ultimos_usuarios = User.objects.order_by('-date_joined')[:10]
    
    # Chat statistics
    total_chats = ConversacionChat.objects.count()
    
    # Music statistics
    total_canciones = Cancion.objects.count()
    total_playlists = Playlist.objects.count()
    total_favoritos = Favorite.objects.count()
    total_reproducciones = 0  # Placeholder - could be tracked separately
    ultimas_canciones = Cancion.objects.order_by('-fecha_agregada')[:5]
    
    # Games statistics
    from apps.app.models import Game, Score, Achievement
    total_juegos = Game.objects.filter(activo=True).count()
    total_puntuaciones = Score.objects.count()
    total_logros = Achievement.objects.count()
    ultimos_juegos = Game.objects.filter(activo=True).order_by('-fecha_creacion')[:5]
    
    # Study statistics
    total_recursos_estudio = StudyResource.objects.count()
    total_sesiones_estudio = StudySession.objects.count()
    total_pomodoros = PomodoroSession.objects.filter(completado=True).count()
    ultimos_recursos = StudyResource.objects.order_by('-fecha_creacion')[:5]
    
    # Weather statistics
    try:
        total_consultas_clima = WeatherCache.objects.count()
    except Exception as e:
        logger.warning(f"Error counting WeatherCache: {str(e)}")
        total_consultas_clima = 0
    
    # Translator statistics
    try:
        total_traducciones = TranslationCache.objects.count()
    except Exception as e:
        logger.warning(f"Error counting TranslationCache: {str(e)}")
        total_traducciones = 0
    
    # Event statistics
    total_eventos = Evento.objects.count()
    
    # Notification statistics
    total_notificaciones = Notificacion.objects.filter(leida=False).count()
    ultimas_notificaciones = Notificacion.objects.order_by('-fecha_creacion')[:5]
    
    # Support statistics
    total_tickets_pendientes = TicketSoporte.objects.filter(estado='abierto').count()
    total_tickets_resueltos = TicketSoporte.objects.filter(estado='resuelto').count()
    
    # Calculate average response time for resolved tickets
    resolved_tickets = TicketSoporte.objects.filter(
        estado='resuelto',
        fecha_resolucion__isnull=False,
        fecha_respuesta__isnull=False
    )
    if resolved_tickets.exists():
        avg_response = resolved_tickets.annotate(
            response_time=Avg('fecha_respuesta' - 'fecha_creacion')
        ).aggregate(avg=Avg('response_time'))['avg']
        tiempo_promedio_respuesta = f"{avg_response.days}d" if avg_response else "N/A"
    else:
        tiempo_promedio_respuesta = "N/A"
    
    # Blog statistics
    total_publicaciones = Post.objects.filter(publicado=True).count()
    total_categorias = Category.objects.count()
    total_comentarios = Comment.objects.count()
    ultimas_publicaciones = Post.objects.filter(publicado=True).order_by('-fecha_publicacion')[:5]
    
    # Suggestions statistics
    ultimas_sugerencias = Sugerencia.objects.order_by('-fecha_creacion')[:5]
    
    # Visitor statistics
    visitantes_hoy = Visitante.objects.filter(fecha_ultima_interaccion__date=today).count()
    visitantes_semana = Visitante.objects.filter(fecha_ultima_interaccion__date__gte=week_ago).count()
    visitantes_mes = Visitante.objects.filter(fecha_ultima_interaccion__date__gte=month_ago).count()
    ultimos_visitantes = Visitante.objects.order_by('-fecha_ultima_interaccion')[:10]
    
    # Security statistics (placeholders for now)
    intentos_fallidos = 0
    sesiones_activas = User.objects.filter(last_login__gte=timezone.now() - timedelta(hours=1)).count()
    
    context = {
        'total_usuarios': total_usuarios,
        'total_chats': total_chats,
        'total_canciones': total_canciones,
        'total_eventos': total_eventos,
        'total_notificaciones': total_notificaciones,
        'total_tickets_pendientes': total_tickets_pendientes,
        'total_tickets_resueltos': total_tickets_resueltos,
        'tiempo_promedio_respuesta': tiempo_promedio_respuesta,
        'total_publicaciones': total_publicaciones,
        'total_categorias': total_categorias,
        'total_comentarios': total_comentarios,
        'total_reproducciones': total_reproducciones,
        'total_playlists': total_playlists,
        'total_favoritos': total_favoritos,
        'total_juegos': total_juegos,
        'total_puntuaciones': total_puntuaciones,
        'total_logros': total_logros,
        'ultimas_canciones': ultimas_canciones,
        'ultimos_juegos': ultimos_juegos,
        'total_recursos_estudio': total_recursos_estudio,
        'total_sesiones_estudio': total_sesiones_estudio,
        'total_pomodoros': total_pomodoros,
        'ultimos_recursos': ultimos_recursos,
        'total_consultas_clima': total_consultas_clima,
        'total_traducciones': total_traducciones,
        'visitantes_hoy': visitantes_hoy,
        'visitantes_semana': visitantes_semana,
        'visitantes_mes': visitantes_mes,
        'ultimos_usuarios': ultimos_usuarios,
        'ultimas_notificaciones': ultimas_notificaciones,
        'ultimas_publicaciones': ultimas_publicaciones,
        'ultimas_sugerencias': ultimas_sugerencias,
        'ultimos_visitantes': ultimos_visitantes,
        'intentos_fallidos': intentos_fallidos,
        'sesiones_activas': sesiones_activas,
    }
    
    return render(request, 'panel_admin.html', context)


@admin_only
def admin_soporte(request):
    """Render admin support panel."""
    return render(request, 'admin_soporte.html')


@admin_only
def admin_sugerencias(request):
    """Render admin suggestions panel."""
    return render(request, 'admin_sugerencias.html')


@admin_only
def panel_admin_email_user(request, user_id):
    """Render email user form."""
    from django.contrib.auth.models import User
    user = User.objects.get(id=user_id)
    return render(request, 'panel_admin_email_user.html', {'user_id': user_id, 'user_target': user})


@require_http_methods(["POST"])
@admin_only
@csrf_exempt
def responder_ticket(request, ticket_id):
    """Respond to a support ticket."""
    try:
        import json
        data = json.loads(request.body)
        respuesta = data.get('respuesta')
        
        # Placeholder for ticket response logic
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error al responder ticket: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["POST"])
@admin_only
@csrf_exempt
def responder_sugerencia(request, sugerencia_id):
    """Respond to a suggestion."""
    try:
        import json
        data = json.loads(request.body)
        respuesta = data.get('respuesta')
        estado = data.get('estado')
        
        # Placeholder for suggestion response logic
        return JsonResponse({'success': True})
    except Exception as e:
        logger.error(f"Error al responder sugerencia: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
@admin_only
def admin_stats_api(request):
    """Get admin statistics."""
    try:
        from django.db.models import Count, Q, Avg
        from apps.app.models import ConversacionChat, Cancion, Playlist
        from apps.eventos.models import Evento
        from apps.notificaciones.models import Notificacion
        from apps.soporte.models import TicketSoporte
        from apps.blog.models import Post, Category, Comment
        from apps.sugerencias.models import Sugerencia, Visitante
        from django.utils import timezone
        from datetime import timedelta
        
        # Calculate dates for filtering
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Gather statistics
        stats = {
            'total_usuarios': User.objects.count(),
            'total_chats': ConversacionChat.objects.count(),
            'total_canciones': Cancion.objects.count(),
            'total_eventos': Evento.objects.count(),
            'total_notificaciones': Notificacion.objects.filter(leida=False).count(),
            'total_tickets_pendientes': TicketSoporte.objects.filter(estado='abierto').count(),
            'total_tickets_resueltos': TicketSoporte.objects.filter(estado='resuelto').count(),
            'total_publicaciones': Post.objects.filter(publicado=True).count(),
            'total_categorias': Category.objects.count(),
            'total_comentarios': Comment.objects.count(),
            'total_reproducciones': 0,  # Placeholder
            'total_playlists': Playlist.objects.count(),
            'visitantes_hoy': Visitante.objects.filter(fecha_ultima_interaccion__date=today).count(),
            'visitantes_semana': Visitante.objects.filter(fecha_ultima_interaccion__date__gte=week_ago).count(),
            'visitantes_mes': Visitante.objects.filter(fecha_ultima_interaccion__date__gte=month_ago).count(),
            'intentos_fallidos': 0,
            'sesiones_activas': User.objects.filter(last_login__gte=timezone.now() - timedelta(hours=1)).count(),
        }
        return JsonResponse(stats)
    except Exception as e:
        logger.error(f"Error al obtener estadísticas: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


def _safe_count(model):
    """Helper function to safely count objects, handling missing tables."""
    try:
        return model.objects.count()
    except Exception:
        return 0


def _safe_count_with_filter(model, filters):
    """Helper function to safely count objects with filters, handling missing tables."""
    try:
        return model.objects.filter(**filters).count()
    except Exception as e:
        logger.warning(f"Error counting {model.__name__} with filters {filters}: {str(e)}")
        return 0


@admin_only
def exportar_reporte_excel(request):
    """Export report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.db.models import Count, Q, Avg
        from apps.app.models import ConversacionChat, Cancion, Playlist, Favorite
        from apps.eventos.models import Evento
        from apps.notificaciones.models import Notificacion
        from apps.soporte.models import TicketSoporte
        from apps.blog.models import Post, Category, Comment
        from apps.sugerencias.models import Sugerencia, Visitante
        from apps.estudio.models import StudyResource, StudySession, PomodoroSession
        from apps.clima.models import WeatherCache
        from apps.traductor.models import TranslationCache
        from apps.app.models import Game, Score, Achievement
        from django.utils import timezone
        from datetime import timedelta
        from django.http import HttpResponse
        import openpyxl
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Administrativo"
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws['A1'] = 'Reporte Administrativo - MiniAmigixV'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:B1')
        
        ws['A2'] = f'Fecha: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws.merge_cells('A2:B2')
        
        # Calculate statistics
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Calculate weekly activity by day
        def get_weekly_activity():
            """Get activity counts for each day of the current week."""
            days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            activity = []
            
            # Get the start of the current week (Monday)
            start_of_week = today - timedelta(days=today.weekday())
            
            for i, day_name in enumerate(days):
                day_date = start_of_week + timedelta(days=i)
                # Count various activities for this day
                day_activity = {
                    'day': day_name,
                    'date': day_date,
                    'chats': _safe_count_with_filter(ConversacionChat, {'fecha_creacion__date': day_date}),
                    'canciones': _safe_count_with_filter(Cancion, {'fecha_agregada__date': day_date}),
                    'eventos': _safe_count_with_filter(Evento, {'fecha_creacion__date': day_date}),
                    'sesiones_estudio': _safe_count_with_filter(StudySession, {'fecha_inicio__date': day_date}),
                }
                activity.append(day_activity)
            
            return activity
        
        weekly_activity = get_weekly_activity()
        
        stats_data = [
            ['Categoría', 'Métrica', 'Valor'],
            ['USUARIOS', 'Total Usuarios', User.objects.count()],
            ['USUARIOS', 'Superusuarios', User.objects.filter(is_superuser=True).count()],
            ['USUARIOS', 'Usuarios Activos (7d)', User.objects.filter(last_login__gte=week_ago).count()],
            ['USUARIOS', 'Sesiones Activas (1h)', User.objects.filter(last_login__gte=timezone.now() - timedelta(hours=1)).count()],
            ['', '', ''],
            ['CHAT IA', 'Total Chats', ConversacionChat.objects.count()],
            ['CHAT IA', 'Chats Últimos 7 días', ConversacionChat.objects.filter(fecha_creacion__gte=week_ago).count()],
            ['', '', ''],
            ['MÚSICA', 'Total Canciones', Cancion.objects.count()],
            ['MÚSICA', 'Total Playlists', Playlist.objects.count()],
            ['MÚSICA', 'Total Favoritos', Favorite.objects.count()],
            ['MÚSICA', 'Canciones Últimos 7 días', Cancion.objects.filter(fecha_agregada__gte=week_ago).count()],
            ['', '', ''],
            ['JUEGOS', 'Total Juegos', Game.objects.filter(activo=True).count()],
            ['JUEGOS', 'Total Puntuaciones', Score.objects.count()],
            ['JUEGOS', 'Total Logros', Achievement.objects.count()],
            ['JUEGOS', 'Puntuaciones Últimos 7 días', _safe_count_with_filter(Score, {'fecha__gte': week_ago})],
            ['', '', ''],
            ['ESTUDIO', 'Total Recursos', StudyResource.objects.count()],
            ['ESTUDIO', 'Total Sesiones', StudySession.objects.count()],
            ['ESTUDIO', 'Pomodoros Completados', PomodoroSession.objects.filter(completado=True).count()],
            ['ESTUDIO', 'Sesiones Últimos 7 días', _safe_count_with_filter(StudySession, {'fecha_inicio__gte': week_ago})],
            ['', '', ''],
            ['EVENTOS', 'Total Eventos', Evento.objects.count()],
            ['EVENTOS', 'Eventos Próximos', Evento.objects.filter(fecha__gte=today).count()],
            ['EVENTOS', 'Eventos Últimos 7 días', _safe_count_with_filter(Evento, {'fecha_creacion__gte': week_ago})],
            ['', '', ''],
            ['CLIMA', 'Consultas de Clima', _safe_count(WeatherCache)],
            ['CLIMA', 'Consultas Últimos 7 días', _safe_count_with_filter(WeatherCache, {'timestamp__gte': week_ago})],
            ['', '', ''],
            ['TRADUCTOR', 'Traducciones', _safe_count(TranslationCache)],
            ['TRADUCTOR', 'Traducciones Últimos 7 días', _safe_count_with_filter(TranslationCache, {'created_at__gte': week_ago})],
            ['', '', ''],
            ['BLOG', 'Publicaciones', Post.objects.filter(publicado=True).count()],
            ['BLOG', 'Categorías', Category.objects.count()],
            ['BLOG', 'Comentarios', Comment.objects.count()],
            ['BLOG', 'Publicaciones Últimos 7 días', Post.objects.filter(publicado=True, fecha_publicacion__gte=week_ago).count()],
            ['', '', ''],
            ['SOPORTE', 'Tickets Pendientes', TicketSoporte.objects.filter(estado='abierto').count()],
            ['SOPORTE', 'Tickets Resueltos', TicketSoporte.objects.filter(estado='resuelto').count()],
            ['SOPORTE', 'Tickets Últimos 7 días', _safe_count_with_filter(TicketSoporte, {'fecha_creacion__gte': week_ago})],
            ['', '', ''],
            ['SUGERENCIAS', 'Total Sugerencias', Sugerencia.objects.count()],
            ['SUGERENCIAS', 'Pendientes', Sugerencia.objects.filter(estado='pendiente').count()],
            ['SUGERENCIAS', 'Aprobadas', Sugerencia.objects.filter(estado='aprobada').count()],
            ['', '', ''],
            ['NOTIFICACIONES', 'Total', Notificacion.objects.count()],
            ['NOTIFICACIONES', 'Sin Leer', Notificacion.objects.filter(leida=False).count()],
            ['NOTIFICACIONES', 'Últimos 7 días', _safe_count_with_filter(Notificacion, {'fecha_creacion__gte': week_ago})],
            ['', '', ''],
            ['VISITANTES', 'Hoy', Visitante.objects.filter(fecha_ultima_interaccion__date=today).count()],
            ['VISITANTES', 'Esta Semana', Visitante.objects.filter(fecha_ultima_interaccion__date__gte=week_ago).count()],
            ['VISITANTES', 'Este Mes', Visitante.objects.filter(fecha_ultima_interaccion__date__gte=month_ago).count()],
            ['VISITANTES', 'Total Único', Visitante.objects.count()],
            ['', '', ''],
            ['ACTIVIDAD SEMANAL', 'Lunes', weekly_activity[0]['chats'] + weekly_activity[0]['canciones'] + weekly_activity[0]['eventos'] + weekly_activity[0]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Martes', weekly_activity[1]['chats'] + weekly_activity[1]['canciones'] + weekly_activity[1]['eventos'] + weekly_activity[1]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Miércoles', weekly_activity[2]['chats'] + weekly_activity[2]['canciones'] + weekly_activity[2]['eventos'] + weekly_activity[2]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Jueves', weekly_activity[3]['chats'] + weekly_activity[3]['canciones'] + weekly_activity[3]['eventos'] + weekly_activity[3]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Viernes', weekly_activity[4]['chats'] + weekly_activity[4]['canciones'] + weekly_activity[4]['eventos'] + weekly_activity[4]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Sábado', weekly_activity[5]['chats'] + weekly_activity[5]['canciones'] + weekly_activity[5]['eventos'] + weekly_activity[5]['sesiones_estudio']],
            ['ACTIVIDAD SEMANAL', 'Domingo', weekly_activity[6]['chats'] + weekly_activity[6]['canciones'] + weekly_activity[6]['eventos'] + weekly_activity[6]['sesiones_estudio']],
        ]
        
        # Write data starting from row 4
        start_row = 4
        for row_idx, row_data in enumerate(stats_data):
            for col_idx, value in enumerate(row_data):
                cell = ws.cell(row=start_row + row_idx, column=col_idx + 1, value=value)
                if row_idx == 0:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_admin_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar reporte Excel: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@admin_only
def exportar_reporte_pdf(request):
    """Export report to PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from django.db.models import Count, Q, Avg
        from apps.app.models import ConversacionChat, Cancion, Playlist, Favorite
        from apps.eventos.models import Evento
        from apps.notificaciones.models import Notificacion
        from apps.soporte.models import TicketSoporte
        from apps.blog.models import Post, Category, Comment
        from apps.sugerencias.models import Sugerencia, Visitante
        from apps.estudio.models import StudyResource, StudySession, PomodoroSession
        from apps.clima.models import WeatherCache
        from apps.traductor.models import TranslationCache
        from apps.app.models import Game, Score, Achievement
        from django.utils import timezone
        from datetime import timedelta
        from django.http import HttpResponse
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=30,
            alignment=1  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # Calculate statistics
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)
        
        # Calculate weekly activity by day
        def get_weekly_activity():
            """Get activity counts for each day of the current week."""
            days = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
            activity = []
            
            # Get the start of the current week (Monday)
            start_of_week = today - timedelta(days=today.weekday())
            
            for i, day_name in enumerate(days):
                day_date = start_of_week + timedelta(days=i)
                # Count various activities for this day
                day_activity = {
                    'day': day_name,
                    'date': day_date,
                    'chats': _safe_count_with_filter(ConversacionChat, {'fecha_creacion__date': day_date}),
                    'canciones': _safe_count_with_filter(Cancion, {'fecha_agregada__date': day_date}),
                    'eventos': _safe_count_with_filter(Evento, {'fecha_creacion__date': day_date}),
                    'sesiones_estudio': _safe_count_with_filter(StudySession, {'fecha_inicio__date': day_date}),
                }
                activity.append(day_activity)
            
            return activity
        
        weekly_activity = get_weekly_activity()
        
        # Build content
        story = []
        
        # Title
        story.append(Paragraph("Reporte Administrativo - MiniAmigixV", title_style))
        story.append(Paragraph(f"Fecha: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))
        
        # Statistics data
        stats_sections = [
            ("USUARIOS", [
                ("Total Usuarios", User.objects.count()),
                ("Superusuarios", User.objects.filter(is_superuser=True).count()),
                ("Usuarios Activos (7d)", User.objects.filter(last_login__gte=week_ago).count()),
                ("Sesiones Activas (1h)", User.objects.filter(last_login__gte=timezone.now() - timedelta(hours=1)).count()),
            ]),
            ("CHAT IA", [
                ("Total Chats", ConversacionChat.objects.count()),
                ("Chats Últimos 7 días", ConversacionChat.objects.filter(fecha_creacion__gte=week_ago).count()),
            ]),
            ("MÚSICA", [
                ("Total Canciones", Cancion.objects.count()),
                ("Total Playlists", Playlist.objects.count()),
                ("Total Favoritos", Favorite.objects.count()),
                ("Canciones Últimos 7 días", Cancion.objects.filter(fecha_agregada__gte=week_ago).count()),
            ]),
            ("JUEGOS", [
                ("Total Juegos", Game.objects.filter(activo=True).count()),
                ("Total Puntuaciones", Score.objects.count()),
                ("Total Logros", Achievement.objects.count()),
                ("Puntuaciones Últimos 7 días", _safe_count_with_filter(Score, {'fecha__gte': week_ago})),
            ]),
            ("ESTUDIO", [
                ("Total Recursos", StudyResource.objects.count()),
                ("Total Sesiones", StudySession.objects.count()),
                ("Pomodoros Completados", PomodoroSession.objects.filter(completado=True).count()),
                ("Sesiones Últimos 7 días", _safe_count_with_filter(StudySession, {'fecha_inicio__gte': week_ago})),
            ]),
            ("EVENTOS", [
                ("Total Eventos", Evento.objects.count()),
                ("Eventos Próximos", Evento.objects.filter(fecha__gte=today).count()),
                ("Eventos Últimos 7 días", _safe_count_with_filter(Evento, {'fecha_creacion__gte': week_ago})),
            ]),
            ("CLIMA", [
                ("Consultas de Clima", _safe_count(WeatherCache)),
                ("Consultas Últimos 7 días", _safe_count_with_filter(WeatherCache, {'timestamp__gte': week_ago})),
            ]),
            ("TRADUCTOR", [
                ("Traducciones", _safe_count(TranslationCache)),
                ("Traducciones Últimos 7 días", _safe_count_with_filter(TranslationCache, {'created_at__gte': week_ago})),
            ]),
            ("BLOG", [
                ("Publicaciones", Post.objects.filter(publicado=True).count()),
                ("Categorías", Category.objects.count()),
                ("Comentarios", Comment.objects.count()),
                ("Publicaciones Últimos 7 días", Post.objects.filter(publicado=True, fecha_publicacion__gte=week_ago).count()),
            ]),
            ("SOPORTE", [
                ("Tickets Pendientes", TicketSoporte.objects.filter(estado='abierto').count()),
                ("Tickets Resueltos", TicketSoporte.objects.filter(estado='resuelto').count()),
                ("Tickets Últimos 7 días", _safe_count_with_filter(TicketSoporte, {'fecha_creacion__gte': week_ago})),
            ]),
            ("SUGERENCIAS", [
                ("Total Sugerencias", Sugerencia.objects.count()),
                ("Pendientes", Sugerencia.objects.filter(estado='pendiente').count()),
                ("Aprobadas", Sugerencia.objects.filter(estado='aprobada').count()),
            ]),
            ("NOTIFICACIONES", [
                ("Total", Notificacion.objects.count()),
                ("Sin Leer", Notificacion.objects.filter(leida=False).count()),
                ("Últimos 7 días", _safe_count_with_filter(Notificacion, {'fecha_creacion__gte': week_ago})),
            ]),
            ("VISITANTES", [
                ("Hoy", Visitante.objects.filter(fecha_ultima_interaccion__date=today).count()),
                ("Esta Semana", Visitante.objects.filter(fecha_ultima_interaccion__date__gte=week_ago).count()),
                ("Este Mes", Visitante.objects.filter(fecha_ultima_interaccion__date__gte=month_ago).count()),
                ("Total Único", Visitante.objects.count()),
            ]),
            ("ACTIVIDAD SEMANAL", [
                ("Lunes", weekly_activity[0]['chats'] + weekly_activity[0]['canciones'] + weekly_activity[0]['eventos'] + weekly_activity[0]['sesiones_estudio']),
                ("Martes", weekly_activity[1]['chats'] + weekly_activity[1]['canciones'] + weekly_activity[1]['eventos'] + weekly_activity[1]['sesiones_estudio']),
                ("Miércoles", weekly_activity[2]['chats'] + weekly_activity[2]['canciones'] + weekly_activity[2]['eventos'] + weekly_activity[2]['sesiones_estudio']),
                ("Jueves", weekly_activity[3]['chats'] + weekly_activity[3]['canciones'] + weekly_activity[3]['eventos'] + weekly_activity[3]['sesiones_estudio']),
                ("Viernes", weekly_activity[4]['chats'] + weekly_activity[4]['canciones'] + weekly_activity[4]['eventos'] + weekly_activity[4]['sesiones_estudio']),
                ("Sábado", weekly_activity[5]['chats'] + weekly_activity[5]['canciones'] + weekly_activity[5]['eventos'] + weekly_activity[5]['sesiones_estudio']),
                ("Domingo", weekly_activity[6]['chats'] + weekly_activity[6]['canciones'] + weekly_activity[6]['eventos'] + weekly_activity[6]['sesiones_estudio']),
            ]),
        ]
        
        # Create tables for each section
        for section_title, metrics in stats_sections:
            story.append(Paragraph(section_title, heading_style))
            
            table_data = [['Métrica', 'Valor']]
            for metric_name, value in metrics:
                table_data.append([metric_name, str(value)])
            
            table = Table(table_data, colWidths=[3 * inch, 2 * inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))
        
        # Build PDF
        doc.build(story)
        
        # Prepare response
        buffer.seek(0)
        response = HttpResponse(
            buffer,
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename=reporte_admin_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        logger.error(f"Error al exportar reporte PDF: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)
